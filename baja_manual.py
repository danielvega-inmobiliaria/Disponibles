"""
Gestion manual de propiedades.

Opciones:
  1. Dar de baja una propiedad vendida (entre listas)
  2. Agregar propiedad manualmente (inversor, no esta en listas)
  3. Reactivar desde historial (estaba vendida y vuelve al mercado)
  4. Ver historial de bajas

Uso: python baja_manual.py
     python baja_manual.py --historial
"""

import openpyxl
import os
import json
from datetime import datetime

# ─── RUTAS ────────────────────────────────────────────────────────────────────
DISPONIBLES_DIR = os.path.dirname(os.path.abspath(__file__))
DESARROLLOS_DIR = os.path.dirname(DISPONIBLES_DIR)
DISP_FILE       = os.path.join(DISPONIBLES_DIR, "DISPONIBLES.xlsx")
HISTORIAL_FILE  = os.path.join(DISPONIBLES_DIR, "vendidas.json")

# ─── COLUMNAS EN DISPONIBLES.xlsx (0-based) ───────────────────────────────────
C_EMPRESA      = 1
# C_PROYECTO eliminado — se usa UBICACIÓN como identificador
C_ZONA         = 3
C_UBICACION    = 4
C_ETAPA        = 5
C_ENTREGA      = 6
C_UNIDAD       = 7
C_TIPO         = 8
C_SUP_CUB      = 10
C_SUP_TOTAL    = 14
C_PRECIO_LISTA = 15
C_PRECIO_CONT  = 16
C_USD_M2       = 17
C_OPORTUNIDAD  = 19
C_NOTAS        = 20
TOTAL_COLS     = 21


# ─── UTILIDADES ───────────────────────────────────────────────────────────────
def parse_price(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v else None
    try:
        return float(str(v).replace("USD","").replace("$","").replace(",","").strip()) or None
    except Exception:
        return None


def fmt_precio(v):
    if v is None:
        return "—"
    return f"U$S {int(v):,}".replace(",",".")


def pedir(prompt, requerido=False):
    while True:
        v = input(f"  {prompt}: ").strip()
        if v or not requerido:
            return v
        print("  Campo requerido.")


def pedir_precio(prompt):
    while True:
        v = input(f"  {prompt} (ej: 95000 o Enter para omitir): ").strip()
        if not v:
            return None
        try:
            return float(v.replace(",","."))
        except Exception:
            print("  Ingresa solo el numero (ej: 95000).")


# ─── LEER / GUARDAR DISPONIBLES ───────────────────────────────────────────────
def read_disponibles():
    wb = openpyxl.load_workbook(DISP_FILE)
    ws = wb.active
    props = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 10:
            continue
        empresa = str(row[C_EMPRESA] or "").strip().upper()
        unidad  = str(row[C_UNIDAD]  or "").strip().upper()
        if not unidad:
            continue
        props.append({
            "fila":        i,
            "empresa":     empresa,
            "unidad":      unidad,
            "zona":        str(row[C_ZONA]       or "").strip(),
            "ubicacion":   str(row[C_UBICACION]  or "").strip(),
            "etapa":       str(row[C_ETAPA]      or "").strip(),
            "entrega":     str(row[C_ENTREGA]    or "").strip(),
            "tipo":        str(row[C_TIPO]       or "").strip(),
            "sup_cub":     parse_price(row[C_SUP_CUB])    if len(row) > C_SUP_CUB    else None,
            "sup_total":   parse_price(row[C_SUP_TOTAL])  if len(row) > C_SUP_TOTAL  else None,
            "precio_lista":parse_price(row[C_PRECIO_LISTA]) if len(row) > C_PRECIO_LISTA else None,
            "precio_cont": parse_price(row[C_PRECIO_CONT])  if len(row) > C_PRECIO_CONT  else None,
            "oportunidad": str(row[C_OPORTUNIDAD] or "").strip() if len(row) > C_OPORTUNIDAD else "",
            "notas":       str(row[C_NOTAS]       or "").strip() if len(row) > C_NOTAS       else "",
        })
    return props, wb, ws


# ─── HISTORIAL ────────────────────────────────────────────────────────────────
def leer_historial():
    if not os.path.exists(HISTORIAL_FILE):
        return []
    try:
        with open(HISTORIAL_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def guardar_historial_file(historial):
    with open(HISTORIAL_FILE, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)


def agregar_a_historial(props_eliminadas, motivo):
    historial = leer_historial()
    ts = datetime.now().isoformat(timespec="seconds")
    for p in props_eliminadas:
        historial.append({
            "fecha_baja":   ts,
            "motivo":       motivo,
            "empresa":      p.get("empresa",""),
            "unidad":       p.get("unidad",""),
            "ubicacion":    p.get("ubicacion",""),
            "zona":         p.get("zona",""),
            "etapa":        p.get("etapa",""),
            "entrega":      p.get("entrega",""),
            "tipo":         p.get("tipo",""),
            "sup_cub":      p.get("sup_cub"),
            "sup_total":    p.get("sup_total"),
            "precio_lista": p.get("precio_lista"),
            "precio_cont":  p.get("precio_cont"),
            "oportunidad":  p.get("oportunidad",""),
            "notas":        p.get("notas",""),
        })
    guardar_historial_file(historial)
    return len(props_eliminadas)


def guardar_backup_y_xlsx(wb):
    import shutil
    ts  = datetime.now().strftime("%Y%m%d_%H%M")
    bak = DISP_FILE.replace(".xlsx", f"_bak_{ts}.xlsx")
    shutil.copy2(DISP_FILE, bak)
    wb.save(DISP_FILE)
    return bak


# ─── 1. BAJA MANUAL ──────────────────────────────────────────────────────────
def baja_manual():
    sep = "─" * 55
    print(f"\n{sep}")
    print("  BAJA MANUAL DE PROPIEDAD VENDIDA")
    print(sep)

    props, wb, ws = read_disponibles()
    print(f"\n  {len(props)} propiedades en DISPONIBLES.xlsx")

    while True:
        print(f"\n  Buscar por:  1-Unidad  2-Direccion  3-Listar todo  0-Volver")
        op = input("  > ").strip()
        if op == "0":
            return
        elif op == "3":
            resultados = props
        elif op == "1":
            busq = input("  Unidad (parcial): ").strip().upper()
            resultados = [p for p in props if busq in p["unidad"]]
        elif op == "2":
            busq = input("  Direccion (parcial): ").strip().upper()
            resultados = [p for p in props
                          if busq in p["ubicacion"].upper()]
        else:
            continue

        if not resultados:
            print("  Sin resultados.")
            continue

        print(f"\n  {'#':<4} {'Empresa':<8} {'Unidad':<14} {'Ub./Proyecto':<25} {'Tipo':<16} {'Precio'}")
        print(f"  {'-'*4} {'-'*8} {'-'*14} {'-'*25} {'-'*16} {'-'*14}")
        for i, p in enumerate(resultados, 1):
            ref = p["ubicacion"][:25]
            print(f"  {i:<4} {p['empresa']:<8} {p['unidad']:<14} {ref:<25} {p['tipo'][:16]:<16} {fmt_precio(p['precio_cont'])}")

        print(f"\n  Numeros a dar de baja (ej: 1,3) o Enter para cancelar:")
        sel = input("  > ").strip()
        if not sel:
            continue
        try:
            nums = [int(x.strip()) for x in sel.split(",")]
            seleccionadas = [resultados[n-1] for n in nums if 1 <= n <= len(resultados)]
        except Exception:
            print("  Entrada invalida.")
            continue

        if not seleccionadas:
            continue

        print(f"\n  DAR DE BAJA:")
        for p in seleccionadas:
            ref = p["ubicacion"][:30]
            print(f"    {p['empresa']} | {p['unidad']:14} | {ref} | {fmt_precio(p['precio_cont'])}")

        motivo = input("\n  Motivo (venta / error / otro): ").strip() or "venta_manual"
        if input("  Confirmar? [s/N]: ").strip().lower() != "s":
            print("  Cancelado.")
            continue

        agregar_a_historial(seleccionadas, motivo)
        for fila in sorted([p["fila"] for p in seleccionadas], reverse=True):
            ws.delete_rows(fila)
        bak = guardar_backup_y_xlsx(wb)

        print(f"\n  OK: {len(seleccionadas)} propiedad(es) dada(s) de baja.")
        print(f"  Historial guardado. Backup: {os.path.basename(bak)}")
        print(f"  Ejecuta opcion 2 del menu para regenerar el HTML.")

        props, wb, ws = read_disponibles()
        if input("\n  Dar de baja otra? [s/N]: ").strip().lower() != "s":
            break


# ─── 2. ALTA MANUAL ──────────────────────────────────────────────────────────
def alta_manual():
    sep = "─" * 55
    print(f"\n{sep}")
    print("  ALTA MANUAL DE PROPIEDAD")
    print(sep)
    print("\n  Completa los datos. Los campos con * son requeridos.")
    print("  (Precio sin simbolos, solo numero. Ej: 95000)\n")

    if not os.path.exists(DISP_FILE):
        print(f"  ERROR: No se encontro {DISP_FILE}")
        return

    empresa   = pedir("Empresa * (M2 / OBRING / OTRO)", requerido=True).upper()
    zona      = pedir("Zona (ej: Fisherton, Centro, Pichincha)")
    ubicacion = pedir("Direccion * (ej: Wilde 455 Bis)", requerido=True)
    etapa     = pedir("Etapa")
    entrega   = pedir("Entrega (ej: 12/2025)")
    unidad    = pedir("Unidad * (ej: 3A, PH, LOTE 5)", requerido=True).upper()
    tipo      = pedir("Tipo * (ej: 2 Dormitorios, Cochera, Local)", requerido=True)
    sup_cub   = pedir_precio("Sup. cubierta m2")
    sup_total = pedir_precio("Sup. total m2")
    p_lista   = pedir_precio("Precio lista U$S")
    p_cont    = pedir_precio("Precio contado U$S")
    oportunidad = pedir("Oportunidad (Alta / Media / Baja)")
    notas     = pedir("Notas")

    # Calcular U$S/m2
    usd_m2 = None
    if p_cont and (sup_total or sup_cub):
        sup = sup_total or sup_cub
        usd_m2 = round(p_cont / sup, 2)

    print(f"\n  RESUMEN:")
    print(f"  Empresa:    {empresa}")
    print(f"  Zona:       {zona}")
    print(f"  Direccion:  {ubicacion}")
    print(f"  Unidad:     {unidad}  |  Tipo: {tipo}")
    print(f"  Precio:     {fmt_precio(p_cont)}  (lista: {fmt_precio(p_lista)})")
    if usd_m2:
        print(f"  U$S/m2:     {usd_m2}")

    if input("\n  Confirmar alta? [s/N]: ").strip().lower() != "s":
        print("  Cancelado.")
        return

    wb = openpyxl.load_workbook(DISP_FILE)
    ws = wb.active

    nueva_fila = [""] * TOTAL_COLS
    nueva_fila[C_EMPRESA]      = empresa
    nueva_fila[C_ZONA]         = zona
    nueva_fila[C_UBICACION]    = ubicacion
    nueva_fila[C_ETAPA]        = etapa
    nueva_fila[C_ENTREGA]      = entrega
    nueva_fila[C_UNIDAD]       = unidad
    nueva_fila[C_TIPO]         = tipo
    nueva_fila[C_SUP_CUB]      = sup_cub
    nueva_fila[C_SUP_TOTAL]    = sup_total
    nueva_fila[C_PRECIO_LISTA] = p_lista
    nueva_fila[C_PRECIO_CONT]  = p_cont
    nueva_fila[C_USD_M2]       = usd_m2
    nueva_fila[C_OPORTUNIDAD]  = oportunidad
    nueva_fila[C_NOTAS]        = notas
    ws.append(nueva_fila)

    bak = guardar_backup_y_xlsx(wb)
    print(f"\n  OK: {empresa} | {unidad} agregada a DISPONIBLES.xlsx")
    print(f"  Backup: {os.path.basename(bak)}")
    print(f"  Ejecuta opcion 2 del menu para regenerar el HTML.")
    print(f"\n  NOTA: Si tenes imagen para este edificio, asegurate de que")
    print(f"  el nombre del archivo empiece con '{ubicacion.upper()} - '")


# ─── 3. REACTIVAR DESDE HISTORIAL ────────────────────────────────────────────
def reactivar_desde_historial():
    sep = "─" * 55
    print(f"\n{sep}")
    print("  REACTIVAR PROPIEDAD DEL HISTORIAL")
    print(sep)

    historial = leer_historial()
    if not historial:
        print("\n  El historial esta vacio.")
        return

    # Filtrar vendidas (las que siguen "dadas de baja", sin fecha_reactivacion)
    disponibles_set = set()
    try:
        props, _, _ = read_disponibles()
        disponibles_set = {(p["empresa"], p["unidad"]) for p in props}
    except Exception:
        pass

    candidatas = [h for h in historial
                  if (h.get("empresa",""), h.get("unidad","")) not in disponibles_set]

    if not candidatas:
        print("\n  No hay propiedades en el historial que no esten ya activas.")
        return

    print(f"\n  Propiedades en historial disponibles para reactivar ({len(candidatas)}):")
    print(f"\n  {'#':<4} {'Fecha baja':<12} {'Empresa':<8} {'Unidad':<14} {'Ub./Proyecto':<25} {'Precio'}")
    print(f"  {'-'*4} {'-'*12} {'-'*8} {'-'*14} {'-'*25} {'-'*14}")
    for i, h in enumerate(candidatas, 1):
        fecha = h.get("fecha_baja","")[:10]
        ref   = h.get("ubicacion","")[:25]
        print(f"  {i:<4} {fecha:<12} {h.get('empresa',''):<8} {h.get('unidad','')[:14]:<14} {ref:<25} {fmt_precio(h.get('precio_cont'))}")

    print(f"\n  Numero a reactivar o Enter para cancelar:")
    sel = input("  > ").strip()
    if not sel:
        return
    try:
        idx = int(sel) - 1
        h = candidatas[idx]
    except Exception:
        print("  Seleccion invalida.")
        return

    print(f"\n  Reactivar: {h.get('empresa')} | {h.get('unidad')} | {h.get('ubicacion')}")
    print(f"  Los datos se restauran del historial. Podes modificar el precio:")
    nuevo_cont  = pedir_precio("Precio contado U$S (Enter para mantener el historico)")
    nuevo_lista = pedir_precio("Precio lista U$S   (Enter para mantener el historico)")

    p_cont  = nuevo_cont  if nuevo_cont  is not None else h.get("precio_cont")
    p_lista = nuevo_lista if nuevo_lista is not None else h.get("precio_lista")

    usd_m2 = None
    sup = h.get("sup_total") or h.get("sup_cub")
    if p_cont and sup:
        usd_m2 = round(p_cont / sup, 2)

    if input("\n  Confirmar reactivacion? [s/N]: ").strip().lower() != "s":
        print("  Cancelado.")
        return

    wb = openpyxl.load_workbook(DISP_FILE)
    ws = wb.active

    nueva_fila = [""] * TOTAL_COLS
    nueva_fila[C_EMPRESA]      = h.get("empresa","")
    nueva_fila[C_ZONA]         = h.get("zona","")
    nueva_fila[C_UBICACION]    = h.get("ubicacion","")
    nueva_fila[C_ETAPA]        = h.get("etapa","")
    nueva_fila[C_ENTREGA]      = h.get("entrega","")
    nueva_fila[C_UNIDAD]       = h.get("unidad","")
    nueva_fila[C_TIPO]         = h.get("tipo","")
    nueva_fila[C_SUP_CUB]      = h.get("sup_cub")
    nueva_fila[C_SUP_TOTAL]    = h.get("sup_total")
    nueva_fila[C_PRECIO_LISTA] = p_lista
    nueva_fila[C_PRECIO_CONT]  = p_cont
    nueva_fila[C_USD_M2]       = usd_m2
    nueva_fila[C_OPORTUNIDAD]  = h.get("oportunidad","")
    nueva_fila[C_NOTAS]        = h.get("notas","")
    ws.append(nueva_fila)

    # Marcar en historial como reactivada
    ts = datetime.now().isoformat(timespec="seconds")
    for entry in historial:
        if (entry.get("empresa") == h.get("empresa") and
            entry.get("unidad")  == h.get("unidad") and
            entry.get("fecha_baja") == h.get("fecha_baja")):
            entry["fecha_reactivacion"] = ts
            break
    guardar_historial_file(historial)

    bak = guardar_backup_y_xlsx(wb)
    print(f"\n  OK: {h.get('empresa')} | {h.get('unidad')} reactivada en DISPONIBLES.xlsx")
    print(f"  Backup: {os.path.basename(bak)}")
    print(f"  Ejecuta opcion 2 del menu para regenerar el HTML.")


# ─── 4. VER HISTORIAL ────────────────────────────────────────────────────────
def _ver_historial():
    historial = leer_historial()
    if not historial:
        print("\n  Historial vacio.")
        return

    activas = [h for h in historial if not h.get("fecha_reactivacion")]
    reactiv = [h for h in historial if h.get("fecha_reactivacion")]

    print(f"\n  HISTORIAL: {len(historial)} registros ({len(activas)} bajas activas, {len(reactiv)} reactivadas)")
    print(f"\n  {'Fecha baja':<12} {'Est.':<5} {'Empresa':<8} {'Unidad':<14} {'Ub./Proyecto':<25} {'Precio'}")
    print(f"  {'-'*12} {'-'*5} {'-'*8} {'-'*14} {'-'*25} {'-'*14}")
    for h in historial[-60:]:
        fecha = h.get("fecha_baja","")[:10]
        est   = "OK" if h.get("fecha_reactivacion") else "BAJA"
        ref   = h.get("ubicacion","")[:25]
        precio = h.get("precio_cont")
        ps = f"U$S {int(precio):,}".replace(",",".") if precio else "—"
        print(f"  {fecha:<12} {est:<5} {h.get('empresa',''):<8} {h.get('unidad','')[:14]:<14} {ref:<25} {ps}")

    if len(historial) > 60:
        print(f"\n  (Mostrando ultimos 60 de {len(historial)} registros)")


# ─── MENÚ PRINCIPAL ───────────────────────────────────────────────────────────
def menu():
    sep = "=" * 55
    while True:
        print(f"\n{sep}")
        print("  GESTION DE PROPIEDADES")
        print(sep)
        print("\n   1. Dar de baja propiedad vendida (entre listas)")
        print("   2. Agregar propiedad manualmente (inversor, no en lista)")
        print("   3. Reactivar desde historial (vuelve al mercado)")
        print("   4. Ver historial de bajas")
        print("   0. Salir")
        op = input("\n  Opcion: ").strip()
        if op == "0":
            break
        elif op == "1":
            baja_manual()
        elif op == "2":
            alta_manual()
        elif op == "3":
            reactivar_desde_historial()
        elif op == "4":
            _ver_historial()
            input("\n  Enter para continuar...")


if __name__ == "__main__":
    import sys
    try:
        if "--historial" in sys.argv:
            _ver_historial()
        else:
            menu()
    except KeyboardInterrupt:
        print("\n\n  Interrumpido.")
    except Exception as e:
        import traceback
        print(f"\n  ERROR INESPERADO: {e}")
        traceback.print_exc()
    input("\nPresiona Enter para cerrar...")
