"""
Genera disponibles_mobile.html a partir de DISPONIBLES.xlsx + imagenes en /IMAGENES
Uso: python generar_html.py
Requiere: pip install openpyxl pillow
"""

import openpyxl, os, json, base64, glob, re
from io import BytesIO

DISPONIBLES_DIR = os.path.dirname(os.path.abspath(__file__))
DESARROLLOS_DIR = os.path.dirname(DISPONIBLES_DIR)
IMAGENES_DIR    = os.path.join(DISPONIBLES_DIR, "IMAGENES")

EDIF = {
    "Entre Rios 430":  {"desc":"PB+13 pisos en el centro de Rosario, a metros del Parque España y el río. 1, 2 y 3 dormitorios con balcón. Terminaciones de primera línea, carpinterías de aluminio con DVH y avance del 90%.",
                        "am":["Piscina","Solarium","Quincho con parrilla","Cocheras","Bicicletero","Espacios verdes"]},
    "Wilde 455 Bis":   {"desc":"Condominio privado de 16 viviendas en Fisherton. Casas de 3 y 4 dormitorios con pileta y parrillero propios, departamentos y local comercial. Acceso controlado, jardines y espacios comunes de categoría.",
                        "am":["Pileta privada por unidad","Parrillero","Jardines","Cocheras","Acceso controlado","Espacios comunes"]},
    "Pellegrini 932":  {"desc":"P·932: torre de 14 pisos sobre Av. Pellegrini, 36 unidades de 1 a 4 dormitorios y penthouse exclusivo de 204 m² con piscina privada y gym. Seguridad VIDEO WALL 24hs, hogar inteligente y UP-Amenities en azotea y piso 1.",
                        "am":["Piscina lounge en azotea","Quincho semicubierto","Solarium","Living fogonero","Gimnasio","Coworking","Zoom Kids","Spa con solarium","Cocheras 3 niveles","Bicicletero","Bauleras"]},
    "Montevideo 961":  {"desc":"M·961: edificio de 9 pisos, 25 unidades de 1 a 3 dormitorios a metros de Av. Pellegrini. Balcones privados, terraza exclusiva 41 m² en piso 8, hogar inteligente con pantalla táctil y cerradura Smart.",
                        "am":["Piscina en azotea","Quincho con parrilla","Fogonero","Gimnasio equipado","Coworking","Playroom","Crucijuegos","Cocheras","Bicicletero","WiFi en comunes"]},
    "Moreno 323":      {"desc":"Edificio de categoría en España-Hospitales. Tipologías de 1, 2 y 3 dormitorios con materiales de primera línea. Full amenities, pisos y revestimientos premium.",
                        "am":["Piscina","Carwash","Juegos infantiles","Plazoleta","Cocheras","Quincho"]},
    "Moreno 1151":     {"desc":"Proyecto residencial en Refinería. Unidades de 1, 2 y 3 dormitorios. Terminaciones modernas y amenities completos.",
                        "am":["Piscina","Quincho","Parrilla","Gimnasio","Cocheras"]},
    "Ituzaingo 1852":  {"desc":"I·1846: 6 pisos, 16 unidades de 1 a 4 dormitorios y dúplex exclusivo en el corazón de Rosario. Balcones vegetados con riego automático, materiales premium, hogar inteligente y seguridad VIDEO WALL 24hs.",
                        "am":["Piscina en azotea con cascada","Quincho con toilette","Solarium","Sector relax","Fitness Center","Sector outdoor","Cocheras con carga eléctrica","Bicicletero","Bauleras","WiFi en comunes"]},
    "3 de Febrero 621":{"desc":"Complejo de dos torres, 51 unidades de 2 y 3 dormitorios en barrio Martín, a pasos del Parque Urquiza y el Monumento. Diseño con vegetación integrada, full amenities y equipamiento de categoría.",
                        "am":["Piscina 7×3 m","Quincho Gourmet con parrilla","Fitness Center","Coworking con patio","Espacios kids","Canchita fútbol/básquet","Rooftop con fogonero","Carwash","Pet Shower","Bauleras","Bicicletero"]},
    "Pueyrredon 1101": {"desc":"Edificio premiado internacionalmente. Pileta y solarium en pisos superiores. Laundry, bicicletero y monitoreo 24hs. Entrega inmediata.",
                        "am":["Pileta","Solarium terraza","Quincho con parrilla","Laundry","Bicicletero","Monitoreo 24hs"]},
    "Olive 954":       {"desc":"Río Arriba — 28 casas panorámicas en altura a metros del Parque Alem. Patios privados orientados al río. Espacios verdes y arquitectura de autor. Entrega inmediata.",
                        "am":["Vista al río","Casas en altura","Espacios verdes","Quincho","Piscina","Parrillero"]},
    "Guemes 2285":     {"desc":"Edificio en Pichincha con vista al río. Pisos exclusivos con terminaciones premium. Pileta, quincho y monitoreo 24hs. Listo para escriturar.",
                        "am":["Pileta","Solarium","Quincho","Monitoreo 24hs","Vista al río","Listo para escriturar"]},
    "Dorrego 635":     {"desc":"Oficina en planta baja con ingreso independiente, patio propio, cochera y baulera incluidos. Ideal para profesionales.",
                        "am":["Ingreso independiente","Cochera","Baulera","Patio propio"]},
}


MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

def fmt_entrega(v):
    """Convierte fecha (datetime o string) a 'Mes Año'. Ej: 'Septiembre 2028'."""
    if v is None:
        return ""
    from datetime import datetime as _dt
    if isinstance(v, _dt):
        return f"{MESES[v.month-1]} {v.year}"
    s = str(v).strip()
    if not s or s in ("None", ""):
        return ""
    # Intentar parsear "YYYY-MM-DD ..." o "YYYY-MM-DD"
    try:
        d = _dt.strptime(s[:10], "%Y-%m-%d")
        return f"{MESES[d.month-1]} {d.year}"
    except Exception:
        return s  # devolver tal cual si no se puede parsear


def safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",","").strip())
    except Exception:
        return None


def encode_img(path, max_w=400, max_h=300, quality=72):
    try:
        from PIL import Image, ImageFile
        ImageFile.LOAD_TRUNCATED_IMAGES = True
        img = Image.open(path)
        img.thumbnail((max_w, max_h), Image.LANCZOS)
        buf = BytesIO()
        img.convert("RGB").save(buf, "JPEG", quality=quality)
        return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()
    except ImportError:
        ext = os.path.splitext(path)[1].lower().lstrip(".")
        mime = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png","gif":"image/gif"}.get(ext,"image/jpeg")
        with open(path,"rb") as f:
            return f"data:{mime};base64," + base64.b64encode(f.read()).decode()
    except Exception as e:
        print(f"    ERROR imagen: {e}")
        return None


def scan_images():
    """
    Escanea IMAGENES/ con patron "EDIFICIO - TIPO.ext"
    Devuelve dict: {PREFIX_UPPER: {"fachada": path, "gallery": [paths]}}
    """
    result = {}
    if not os.path.exists(IMAGENES_DIR):
        print(f"  Carpeta IMAGENES no encontrada en {IMAGENES_DIR}")
        return result
    for fname in sorted(os.listdir(IMAGENES_DIR)):
        if " - " not in fname:
            continue
        ext = os.path.splitext(fname)[1].lower()
        if ext not in (".jpg",".jpeg",".png",".gif",".webp",".tif",".tiff"):
            continue
        prefix, resto = fname.split(" - ", 1)
        tipo = os.path.splitext(resto)[0].upper()
        key  = prefix.upper()
        path = os.path.join(IMAGENES_DIR, fname)
        if key not in result:
            result[key] = {"fachada": None, "gallery": []}
        if tipo == "FACHADA":
            result[key]["fachada"] = path
        else:
            result[key]["gallery"].append(path)
    return result


def normalize(s):
    """Quita acentos, normaliza espacios y pone en mayusculas."""
    import unicodedata
    nfkd = unicodedata.normalize('NFKD', str(s))
    sin_tilde = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return ' '.join(sin_tilde.upper().split())

def best_match(prefix_upper, direcciones):
    """Busca la direccion del data que mejor coincide con el prefix del archivo."""
    norm_prefix = normalize(prefix_upper)
    # 1. exacto normalizado
    for d in direcciones:
        if normalize(d) == norm_prefix:
            return d
    # 2. por numero principal
    nums = re.findall(r'\d+', norm_prefix)
    if nums:
        for d in direcciones:
            d_norm = normalize(d)
            d_nums = re.findall(r'\d+', d_norm)
            if d_nums and nums[0] == d_nums[0]:
                words = [w for w in norm_prefix.split() if len(w) > 3 and not w.isdigit()]
                if any(w in d_norm for w in words):
                    return d
    return None


def read_disponibles(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    props = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if len(row) < 10:
            continue
        empresa = str(row[1] or "").strip()
        unidad  = str(row[7] or "").strip()
        if not unidad and not empresa:
            continue
        props.append({
            "id":          i + 1,
            "empresa":     empresa,
            "zona":        str(row[3] or "").strip(),  # col 3: ZONA
            "direccion":   str(row[4] or "").strip(),  # col 4: UBICACION
            "etapa":       str(row[5] or "").strip(),  # col 5: ETAPA
            "entrega":     fmt_entrega(row[6]),         # col 6: F. ENTREGA → "Mes Año"
            "unidad":      unidad,
            "tipo":        str(row[8] or "").strip(),
            "descripcion": str(row[9] or "").strip() if len(row) > 9 else "",
            "sup_cub":     row[10] if len(row) > 10 else None,
            "semi_cub":    row[11] if len(row) > 11 else None,
            "sup_patio":   row[12] if len(row) > 12 else None,
            "sup_total":   row[14] if len(row) > 14 else None,
            "precio_lista": safe_float(row[15]) if len(row) > 15 else None,
            "precio_cont":  safe_float(row[16]) if len(row) > 16 else None,
            "usd_m2":       safe_float(row[17]) if len(row) > 17 else None,
            "oportunidad":  str(row[19] or "").strip() if len(row) > 19 else "",  # col 19: OPORTUNIDAD
            "notas":        str(row[20] or "").strip() if len(row) > 20 else "",  # col 20: NOTAS
        })
    return props


def find_disp_file():
    base = os.path.join(DISPONIBLES_DIR, "DISPONIBLES.xlsx")
    print(f"  Usando: DISPONIBLES.xlsx")
    return base


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Disponibles</title>
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
:root{--pri:#1a3c5e;--acc:#e8a020;--bg:#f0f3f8;--wh:#fff;--txt:#1e2b3a;--mu:#6b7a8d;--br:#dde3ec;--ok:#1d7a45;--tbg:#eef2f8;--r:12px}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--txt)}
#vLista{display:block}#vDetalle{display:none}
.hdr{background:var(--pri);color:#fff;padding:14px 14px 10px}
.hdr h1{font-size:17px;font-weight:700}
.hdr p{font-size:11px;opacity:.6;margin-top:2px}
.bdg{display:inline-block;background:var(--acc);color:#fff;border-radius:20px;padding:1px 8px;font-size:12px;font-weight:800;margin-left:6px;vertical-align:middle}
.fil{background:var(--wh);padding:10px 12px;border-bottom:1px solid var(--br);display:flex;flex-direction:column;gap:8px}
.frow{display:flex;gap:7px;align-items:center}
select,input[type=number]{flex:1;border:1.5px solid var(--br);border-radius:9px;padding:9px 10px;font-size:15px;color:var(--txt);background:var(--bg);outline:none;-webkit-appearance:none;min-width:0}
#brst{background:none;border:1.5px solid var(--br);border-radius:9px;padding:9px 14px;font-size:14px;color:var(--mu);cursor:pointer;flex-shrink:0;display:block}
.fopp{display:flex;gap:6px;align-items:center}
.fopp span{font-size:12px;color:var(--mu);white-space:nowrap;flex-shrink:0}
.fbtn{border:1.5px solid var(--br);border-radius:20px;padding:5px 12px;font-size:12px;font-weight:700;cursor:pointer;background:none;color:var(--mu)}
.fbtn.act-alta{background:#fef2f2;color:#991b1b;border-color:#fca5a5}
.fbtn.act-med{background:#fef3e2;color:#92400e;border-color:#fcd34d}
.fbtn.act-baj{background:#f0fdf4;color:#166534;border-color:#86efac}
.list{padding:10px 10px 80px}
a.card{display:flex;align-items:stretch;width:100%;text-align:left;background:var(--wh);border-radius:var(--r);margin-bottom:9px;box-shadow:0 1px 5px rgba(0,0,0,.08);overflow:hidden;border:none;border-left:4px solid #dde3ec;cursor:pointer;padding:0;text-decoration:none;color:var(--txt)}
a.card:active{opacity:.75}
a.card.cd{border-left-color:#3b82f6}a.card.cc{border-left-color:#94a3b8}
a.card.cb{border-left-color:#cbd5e1}a.card.co{border-left-color:var(--acc)}
.cthumb{width:86px;min-width:86px;background:linear-gradient(135deg,#1a3c5e,#2c6ea0);overflow:hidden;flex-shrink:0;position:relative}
.cthumb img{position:absolute;inset:0;width:100%;height:100%;object-fit:cover}
.cph{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;font-size:26px;opacity:.3;color:#fff}
.cbd{padding:10px 12px;flex:1;min-width:0;display:flex;flex-direction:column;gap:5px}
.crow{display:flex;justify-content:space-between;align-items:flex-start;gap:6px}
.cname{font-size:14px;font-weight:700;line-height:1.25}
.caddr{font-size:11px;color:var(--mu);margin-top:1px}
.cprice{text-align:right;flex-shrink:0}
.pco{font-size:14px;font-weight:800;color:var(--ok);white-space:nowrap}
.pli{font-size:10px;color:var(--mu);text-decoration:line-through}
.tags{display:flex;flex-wrap:wrap;gap:4px}
.tag{background:var(--tbg);color:var(--pri);border-radius:5px;padding:2px 7px;font-size:10px;font-weight:700}
.tz{background:#fef3e2;color:#92400e}.ti{background:#fef2f2;color:#991b1b}
.dhdr{background:var(--pri);color:#fff;display:flex;align-items:center}
#btnb{background:rgba(255,255,255,.15);border:none;padding:14px 16px;color:#fff;font-size:22px;cursor:pointer;flex-shrink:0;font-family:inherit;display:block}
#btnb:active{background:rgba(255,255,255,.3)}
.dtitle{padding:12px 14px 12px 0;flex:1;min-width:0}
.dtitle h2{font-size:15px;font-weight:800;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.dtitle p{font-size:11px;opacity:.7;margin-top:2px}
.hero{height:210px;display:flex;overflow:hidden}
.hero-l{flex:1;background:linear-gradient(135deg,#1a3c5e,#2c6ea0);overflow:hidden;position:relative}
.hero-l img{width:100%;height:100%;object-fit:cover;display:block}
.hero-l .hph{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;font-size:40px;opacity:.25;color:#fff}
.hero-r{width:50%;display:grid;grid-template-columns:1fr 1fr;grid-template-rows:1fr 1fr;gap:2px;background:#0d2035;flex-shrink:0}
.hcell{background:#1a3c5e;display:flex;align-items:center;justify-content:center;overflow:hidden}
.hcell img{width:100%;height:100%;object-fit:cover;display:block}
.hcph{font-size:18px;opacity:.2;color:#fff}
.dbd{padding:13px 13px 80px}
.pbox{background:var(--wh);border-radius:var(--r);padding:14px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,.07)}
.lbl{font-size:10px;color:var(--mu);text-transform:uppercase;letter-spacing:.6px;font-weight:700}
.pco2{font-size:26px;font-weight:900;color:var(--ok);margin:3px 0 1px}
.pli2{font-size:13px;color:var(--mu)}
.pm2{font-size:12px;color:var(--pri);font-weight:700;margin-top:5px}
a.btnmap{display:flex;align-items:center;justify-content:center;gap:8px;background:var(--pri);color:#fff;border-radius:var(--r);padding:13px;text-decoration:none;font-size:14px;font-weight:700;margin-bottom:10px;box-shadow:0 2px 8px rgba(26,60,94,.3)}
.sec{background:var(--wh);border-radius:var(--r);padding:14px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,.07)}
.sec h3{font-size:11px;font-weight:800;color:var(--mu);text-transform:uppercase;letter-spacing:.6px;margin-bottom:10px}
.desc{font-size:14px;line-height:1.65}
.dr{display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid var(--br);gap:10px}
.dr:last-child{border-bottom:none;padding-bottom:0}
.dk{font-size:13px;color:var(--mu);flex-shrink:0}.dv{font-size:13px;font-weight:600;text-align:right}
.agrid{display:grid;grid-template-columns:1fr 1fr;gap:7px}
.am{display:flex;align-items:center;gap:7px;background:var(--tbg);border-radius:9px;padding:8px 10px;font-size:12px;font-weight:600;color:var(--pri)}
.nota{background:#fffbf0;border:1.5px solid #fcd34d;border-radius:9px;padding:10px 12px;font-size:13px;color:#78350f;line-height:1.5;margin-top:3px}
.b2{display:inline-block;border-radius:7px;padding:3px 9px;font-size:12px;font-weight:700}
.inv{background:#fef2f2;color:#991b1b}.alta{background:#fef2f2;color:#991b1b}
.med{background:#fef3e2;color:#92400e}.baj{background:#f0fdf4;color:#166534}
#dbg{background:#fffbf0;border:2px solid #fcd34d;padding:10px 12px;font-size:12px;color:#78350f;display:none}
</style>
</head>
<body>
<div id="dbg"></div>
<div id="vLista">
  <div class="hdr">
    <h1>Propiedades Disponibles <span class="bdg" id="cnt">cargando...</span></h1>
    <p>M2 Desarrollos &middot; Obring &mdash; Mayo 2026</p>
  </div>
  <div class="fil">
    <div class="frow">
      <select id="ft">
        <option value="">Todos los tipos</option>
        <option>1 Dormitorio</option><option>2 Dormitorios</option>
        <option>3 Dormitorios</option><option>4 Dormitorios</option>
        <option>Monoambiente</option><option>Estudio</option><option>Loft</option>
        <option>Cochera</option><option>Baulera</option>
        <option>Oficina</option><option>Local</option><option>Local comercial</option><option>Terreno</option>
      </select>
      <select id="fz">
        <option value="">Todas las zonas</option>
        <option>Centro</option><option>Fisherton</option><option>Pichincha</option>
        <option>Arroyito</option><option>Lourdes</option><option>Abasto</option>
        <option>Agote</option><option>Refineria</option><option>Esp y Hospitales</option>
      </select>
    </div>
    <div class="frow">
      <span style="color:var(--mu);font-size:13px;white-space:nowrap;flex-shrink:0">U$S</span>
      <input type="number" id="fmn" placeholder="Min" min="0" step="5000">
      <span style="color:var(--mu);font-size:13px;white-space:nowrap;flex-shrink:0">&ndash;</span>
      <input type="number" id="fmx" placeholder="Max" min="0" step="5000">
      <select id="fe"><option value="">Ambas</option><option>M2</option><option>OBRING</option></select>
      <button id="brst" type="button">X</button>
    </div>
    <div class="fopp">
      <span>Oportunidad:</span>
      <button class="fbtn" id="fo-alta" type="button" data-opp="Alta">Alta</button>
      <button class="fbtn" id="fo-med"  type="button" data-opp="Media">Media</button>
      <button class="fbtn" id="fo-baj"  type="button" data-opp="Baja">Baja</button>
    </div>
  </div>
  <div class="list" id="list"></div>
</div>
<div id="vDetalle">
  <div class="dhdr">
    <button id="btnb" type="button">&#8592;</button>
    <div class="dtitle"><h2 id="dtit">&mdash;</h2><p id="dsub">&mdash;</p></div>
  </div>
  <div class="hero">
    <div class="hero-l">
      <img id="dimg" src="" alt="" style="display:none">
      <div class="hph" id="dph">&#127970;</div>
    </div>
    <div class="hero-r" id="hgrid">
      <div class="hcell"><div class="hcph">&#128247;</div></div>
      <div class="hcell"><div class="hcph">&#128247;</div></div>
      <div class="hcell"><div class="hcph">&#128247;</div></div>
      <div class="hcell"><div class="hcph">&#128247;</div></div>
    </div>
  </div>
  <div class="dbd" id="dbd"></div>
</div>
<script>
(function(){
try{
var DATA=__DATA__;
var _savedScroll=0;
var LIMGS=__LIMGS__;
var GALLERY=__GALLERY__;
var RIMGS=__RIMGS__;
var EDIF=__EDIF__;
function getImg(d){return LIMGS[d]||RIMGS[d]||null;}
function getEd(d){return EDIF[d]||null;}
function fP(n){if(n==null)return '--';return 'U$S '+Math.round(n).toString().replace(/\B(?=(\d{3})+(?!\d))/g,'.');}
function fN(n,s){if(n==null||n===''||n==='null')return null;return Number(n).toString().replace(/\B(?=(\d{3})+(?!\d))/g,'.')+(s||'');}
function tc(t){t=(t||'').toLowerCase();if(/dorm|mono|estudio|loft/.test(t))return 'cd';if(/cochera/.test(t))return 'cc';if(/baulera/.test(t))return 'cb';return 'co';}
var _activeOpp='';
function setOpp(v){
  _activeOpp=(_activeOpp===v)?'':v;
  ['alta','med','baj'].forEach(function(k){document.getElementById('fo-'+k).className='fbtn';});
  if(_activeOpp){var km={'Alta':'alta','Media':'med','Baja':'baj'};document.getElementById('fo-'+km[_activeOpp]).className='fbtn act-'+km[_activeOpp];}
  go();
}
function go(){
  try{
  var tipo=document.getElementById('ft').value,zona=document.getElementById('fz').value,emp=document.getElementById('fe').value;
  var mn=parseFloat(document.getElementById('fmn').value)||0,mx=parseFloat(document.getElementById('fmx').value)||Infinity;
  var f=[];
  for(var i=0;i<DATA.length;i++){var p=DATA[i];if(tipo&&p.tipo!==tipo)continue;if(zona&&p.zona!==zona)continue;if(emp&&p.empresa!==emp)continue;var pr=p.precio_cont||p.precio_lista||0;if(pr<mn||pr>mx)continue;if(_activeOpp&&p.oportunidad!==_activeOpp)continue;f.push(p);}
  document.getElementById('cnt').textContent=f.length;
  if(!f.length){document.getElementById('list').innerHTML='<div style="text-align:center;padding:60px 20px;color:#6b7a8d"><p style="font-size:15px;font-weight:600">Sin resultados</p></div>';return;}
  var h='';
  for(var i=0;i<f.length;i++){
    var p=f[i],pr=p.precio_cont||p.precio_lista,tl=(p.precio_lista&&p.precio_cont&&p.precio_lista!==p.precio_cont),img=getImg(p.direccion);
    var thumb=img?('<div class="cthumb"><img src="'+img+'" alt="" loading="lazy"></div>'):'<div class="cthumb"><div class="cph">&#127970;</div></div>';
    var tags='<span class="tag">'+p.tipo+'</span>';
    if(p.zona)tags+=' <span class="tag tz">'+p.zona+'</span>';
    var sup=p.sup_total||p.sup_cub;if(sup)tags+=' <span class="tag">'+sup+'m2</span>';
    if(p.entrega)tags+=' <span class="tag">Entrega '+p.entrega+'</span>';
    if(p.notas==='INVERSOR')tags+=' <span class="tag ti">INVERSOR</span>';
    h+='<a class="card '+tc(p.tipo)+'" href="javascript:void(0)" data-id="'+p.id+'">'+thumb+'<div class="cbd"><div class="crow"><div><div class="cname">'+p.empresa+' &middot; '+(p.unidad||p.direccion)+'</div><div class="caddr">&#128205; '+p.direccion+'</div></div><div class="cprice"><div class="pco">'+fP(pr)+'</div>'+(tl?'<div class="pli">'+fP(p.precio_lista)+'</div>':'')+'</div></div><div class="tags">'+tags+'</div></div></a>';
  }
  document.getElementById('list').innerHTML=h;
  }catch(e2){dbg('go: '+e2.message);}
}
function dbg(msg){var el=document.getElementById('dbg');el.style.display='block';el.textContent='ERROR: '+msg;}
function rst(){['ft','fz','fe'].forEach(function(id){document.getElementById(id).value='';});['fmn','fmx'].forEach(function(id){document.getElementById(id).value='';});_activeOpp='';['alta','med','baj'].forEach(function(k){document.getElementById('fo-'+k).className='fbtn';});go();}
function dr(k,v){if(!v||v==='--'||v==='null'||!(v+'').trim())return '';return '<div class="dr"><span class="dk">'+k+'</span><span class="dv">'+v+'</span></div>';}
function openD(id){
  try{
  var pid=parseInt(id,10),p=null;
  for(var i=0;i<DATA.length;i++){if(DATA[i].id===pid){p=DATA[i];break;}}
  if(!p)return;
  var dir=p.direccion,img=getImg(dir),ed=getEd(dir),gal=GALLERY[dir]||[];
  var ie=document.getElementById('dimg'),ph=document.getElementById('dph');
  if(img){ie.src=img;ie.style.display='block';ph.style.display='none';}else{ie.src='';ie.style.display='none';ph.style.display='flex';}
  var gc='';for(var g=0;g<4;g++){gc+=gal[g]?('<div class="hcell"><img src="'+gal[g]+'" alt="" loading="lazy"></div>'):'<div class="hcell"><div class="hcph">&#128247;</div></div>';}
  document.getElementById('hgrid').innerHTML=gc;
  document.getElementById('dtit').textContent=p.empresa+' - '+(p.unidad||p.direccion);
  document.getElementById('dsub').textContent=p.tipo+' - '+dir+(p.zona?' - '+p.zona:'');
  var mu='https://maps.google.com/maps?q='+encodeURIComponent(dir+', Rosario, Argentina');
  var h='<div class="pbox"><div class="lbl">Precio contado</div><div class="pco2">'+fP(p.precio_cont||p.precio_lista)+'</div>'+(p.precio_lista&&p.precio_cont&&p.precio_lista!==p.precio_cont?'<div class="pli2">Lista: '+fP(p.precio_lista)+'</div>':'')+(p.usd_m2?'<div class="pm2">'+Math.round(p.usd_m2).toString().replace(/\B(?=(\d{3})+(?!\d))/g,'.')+' U$S/m2</div>':'')+'</div>';
  h+='<a class="btnmap" href="'+mu+'" target="_blank" rel="noopener">&#128205; Ver en mapa - '+dir+'</a>';
  if(ed&&ed.desc)h+='<div class="sec"><h3>Sobre el edificio</h3><p class="desc">'+ed.desc+'</p></div>';
  if(ed&&ed.am&&ed.am.length){var ams='';for(var a=0;a<ed.am.length;a++){ams+='<div class="am">&#10003; '+ed.am[a]+'</div>';}h+='<div class="sec"><h3>Amenities</h3><div class="agrid">'+ams+'</div></div>';}
  h+='<div class="sec"><h3>Unidad</h3>'+dr('Tipo',p.tipo)+dr('Empresa',p.empresa)+dr('Etapa',p.etapa)+dr('Entrega',p.entrega)+'</div>';
  var s='';if(p.sup_cub)s+=dr('Sup. Cubierta',fN(p.sup_cub,' m2'));if(p.semi_cub)s+=dr('Semi Cubierta',fN(p.semi_cub,' m2'));if(p.sup_patio)s+=dr('Patio / Balcón',fN(p.sup_patio,' m2'));if(p.sup_total)s+=dr('Total',fN(p.sup_total,' m2'));
  if(s)h+='<div class="sec"><h3>Superficies</h3>'+s+'</div>';
  if(p.descripcion&&(p.descripcion+'').trim())h+='<div class="sec"><h3>Descripción</h3><p class="desc">'+p.descripcion+'</p></div>';
  var ob='',om={'Alta':'alta','Media':'med','Baja':'baj'};
  if(p.oportunidad)ob+=dr('Oportunidad','<span class="b2 '+(om[p.oportunidad]||'med')+'">'+p.oportunidad+'</span>');
  if(p.notas==='INVERSOR')ob+=dr('Perfil','<span class="b2 inv">INVERSOR</span>');
  else if(p.notas&&p.notas!=='null'&&(p.notas+'').trim())ob+='<div class="nota">'+p.notas+'</div>';
  if(ob)h+='<div class="sec"><h3>Observaciones</h3>'+ob+'</div>';
  document.getElementById('dbd').innerHTML=h;
  _savedScroll=window.scrollY||window.pageYOffset||0;
  document.getElementById('vLista').style.display='none';
  document.getElementById('vDetalle').style.display='block';
  window.scrollTo(0,0);
  }catch(e3){dbg('openD: '+e3.message);}
}
document.getElementById('ft').addEventListener('change',go);
document.getElementById('fz').addEventListener('change',go);
document.getElementById('fe').addEventListener('change',go);
document.getElementById('fmn').addEventListener('input',go);
document.getElementById('fmx').addEventListener('input',go);
document.getElementById('brst').addEventListener('click',rst);
document.getElementById('fo-alta').addEventListener('click',function(){setOpp('Alta');});
document.getElementById('fo-med').addEventListener('click',function(){setOpp('Media');});
document.getElementById('fo-baj').addEventListener('click',function(){setOpp('Baja');});
document.getElementById('btnb').addEventListener('click',function(){document.getElementById('vDetalle').style.display='none';document.getElementById('vLista').style.display='block';window.scrollTo(0,_savedScroll);});
document.getElementById('list').addEventListener('click',function(e){var el=e.target;while(el&&el!==this){if(el.dataset&&el.dataset.id){openD(el.dataset.id);return;}el=el.parentElement;}});
go();
}catch(e){var d=document.getElementById('dbg');if(d){d.style.display='block';d.textContent='ERROR: '+e.message;}}
})();
</script>
</body>
</html>"""

RIMGS = {
    "Entre Rios 430":  "https://m2desarrollos.com.ar/wp-content/uploads/2026/01/Entre-Rios-430-28.webp",
    "Wilde 455 Bis":   "https://m2desarrollos.com.ar/wp-content/uploads/2026/01/Barrio-Entre-Rios-430-1.webp",
    "Pueyrredon 1101": "https://m2desarrollos.com.ar/wp-content/uploads/2026/01/BARRIO-PELLEGRINI-932-5.webp",
    "Olive 954":       "https://m2ddesarrollos.com.ar/wp-content/uploads/2026/01/Barrio-Entre-Rios-430-3.webp",
    "Guemes 2285":     "https://m2desarrollos.com.ar/wp-content/uploads/2026/01/Entre-Rios-430-26.webp",
}


def main():
    print("=" * 55)
    print("  GENERADOR HTML - DISPONIBLES")
    print("=" * 55)

    disp_file = find_disp_file()
    if not os.path.exists(disp_file):
        print(f"ERROR: No se encontro {disp_file}")
        return

    print(f"\nLeyendo propiedades de {os.path.basename(disp_file)}...")

    wb_dbg = openpyxl.load_workbook(disp_file, data_only=True)
    ws_dbg = wb_dbg.active
    rows_dbg = list(ws_dbg.iter_rows(min_row=1, max_row=2, values_only=True))
    print("  HEADERS:")
    for ci, val in enumerate(rows_dbg[0]):
        print(f"    col {ci}: {val}")
    if len(rows_dbg) > 1:
        print("  FILA 1:")
        for ci, val in enumerate(rows_dbg[1]):
            print(f"    col {ci}: {val}")

    props = read_disponibles(disp_file)
    print(f"  {len(props)} propiedades")

    direcciones = list(set(p["direccion"] for p in props if p["direccion"]))
    print(f"  Ubicaciones unicas: {len(direcciones)}")
    if direcciones:
        print(f"  Muestra: {direcciones[:5]}")
    else:
        print("  AVISO: No se encontraron ubicaciones. Verificar columnas del xlsx.")

    print("\nEscaneando imagenes en /IMAGENES ...")
    img_data = scan_images()
    print(f"  {len(img_data)} edificios con imagenes")

    limgs   = {}
    gallery = {}

    for prefix_upper, imgs in img_data.items():
        match = best_match(prefix_upper, direcciones)
        label = match if match else f"?? {prefix_upper}"

        if imgs["fachada"]:
            enc = encode_img(imgs["fachada"], 400, 300, 72)
            if enc:
                if match:
                    limgs[match] = enc
                print(f"  FACHADA  {label} ({len(enc)//1024}KB)")
            else:
                print(f"  ERROR    {label} fachada")
        else:
            print(f"  SIN FACHADA  {label}")

        gal = []
        for gpath in imgs["gallery"][:4]:
            enc = encode_img(gpath, 320, 240, 68)
            if enc:
                gal.append(enc)
        if gal and match:
            gallery[match] = gal
            print(f"  GALERIA  {label} ({len(gal)} imgs)")

    print(f"\n  Fachadas embebidas:  {len(limgs)}")
    print(f"  Galerias embebidas:  {len(gallery)}")

    sin_match = [k for k in img_data if not best_match(k, direcciones)]
    if sin_match:
        print(f"\n  AVISO - Sin match en DISPONIBLES.xlsx:")
        for k in sin_match:
            print(f"    {k}")

    print("\nGenerando HTML...")
    data_js    = json.dumps(props,   ensure_ascii=False)
    limgs_js   = json.dumps(limgs,   ensure_ascii=False)
    gallery_js = json.dumps(gallery, ensure_ascii=False)
    rimgs_js   = json.dumps(RIMGS,   ensure_ascii=False)
    edif_js    = json.dumps(EDIF,    ensure_ascii=False)

    html = (HTML_TEMPLATE
            .replace("__DATA__",    data_js)
            .replace("__LIMGS__",   limgs_js)
            .replace("__GALLERY__", gallery_js)
            .replace("__RIMGS__",   rimgs_js)
            .replace("__EDIF__",    edif_js))

    for fname in ["disponibles_mobile.html", "index.html"]:
        out = os.path.join(DISPONIBLES_DIR, fname)
        with open(out, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"  OK  {fname} ({len(html)//1024} KB)")

    print("\n" + "=" * 55)
    print("  LISTO.")
    if sin_match:
        print(f"  OJO: {len(sin_match)} edificios sin match en DISPONIBLES.xlsx")
        print("  Verificar que las direcciones coincidan exactamente.")
    print("=" * 55)


if __name__ == "__main__":
    main()
ue las direcciones coincidan exactamente.")
    print("=" * 55)


if __name__ == "__main__":
    main()
