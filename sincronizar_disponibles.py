"""
Sincronizar DISPONIBLES.xlsx con listas nuevas de M2 y Obring.

Detecta:
  - Cambios de precio en unidades existentes
  - Unidades que desaparecen de las listas (posibles ventas)
  - Unidades nuevas que aparecen en las listas

Uso: python sincronizar_disponibles.py
"""

import openpyxl
import os
import glob
import json
import re
from datetime import datetime
from copy import copy

# ─── RUTAS ────────────────────────────────────────────────────────────────────
DISPONIBLES_DIR = os.path.dirname(os.path.abspath(__file__))
DESARROLLOS_DIR = os.path.dirname(DISPONIBLES_DIR)
DISP_FILE       = os.path.join(DISPONIBLES_DIR, "DISPONIBLES.xlsx")
M2_LISTAS_DIR   = os.path.join(DESARROLLOS_DIR, "M2 DESARROLLOS", "LISTAS DE PRECIO")
OBRING_DIR      = os.path.join(DESARROLLOS_DIR, "OBRING")
EXCLUIR_FILE    = os.path.join(DISPONIBLES_DIR, "excluir.json")

# ─── COLUMNAS EN DISPONIBLES.xlsx (0-based) ───────────────────────────────────
C_EMPRESA      = 1
# C_PROYECTO eliminado — se usa UBICACIÓN como identificador
C_ZONA         = 3
C_UBICACION    = 4
C_ETAPA        = 5
C_ENTREGA      = 6
C_UNIDAD       = 7
C_TIPO         = 8
C_SUP_CUB      = 10
C_SUP_TOTAL    = 14
C_PRECIO_LISTA = 15
C_PRECIO_CONT  = 16
C_USD_M2       = 17


# ─── UTILIDADES ───────────────────────────────────────────────────────────────
def parse_price(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v else None
    s = str(v).replace("USD","").replace("$","").replace(",","").strip()
    try:
        return float(s) or None
    except Exception:
        return None


def find_latest_file(directory, pattern="*.xlsx"):
    """Busca el .xlsx más reciente en un directorio."""
    files = sorted(glob.glob(os.path.join(directory, pattern)), reverse=True)
    return files[0] if files else None


def find_col(headers, *candidates):
    """
    Busca la primera columna cuyo encabezado contenga alguno de los candidatos.
    Devuelve el índice o None. Reporta si no encuentra.
    """
    h_upper = [str(h or "").strip().upper() for h in headers]
    for cand in candidates:
        cand_u = cand.upper()
        for i, h in enumerate(h_upper):
            if cand_u in h:
                return i
    return None


# ─── LECTURA M2 ───────────────────────────────────────────────────────────────
def read_m2(filepath):
    """
    Lee todas las hojas de la lista M2.
    Devuelve dict: {EMPRESA_UNIDAD_KEY: {"lista": float, "cont": float, "hoja": str}}
    donde KEY = "UNIDAD_UPPER"
    Detecta columnas dinámicamente. Avisa si falta alguna.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    resultado = {}
    alertas = []

    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None

        # Buscar fila de encabezado
        for i, row in enumerate(rows):
            cells = [str(c or "").strip().upper() for c in row]
            if "UNIDAD" in cells or "TIPO" in cells:
                header_idx = i
                break

        if header_idx is None:
            alertas.append(f"  [M2] Hoja '{sh}': no se encontró fila de encabezado (UNIDAD). Saltando.")
            continue

        headers = rows[header_idx]
        col_u    = find_col(headers, "UNIDAD", "TIPO")
        col_lst  = find_col(headers, "LISTA", "PRECIO LISTA", "P.LISTA")
        col_cont = find_col(headers, "CONTADO", "PRECIO CONTADO", "P.CONTADO", "CONT")

        faltantes = []
        if col_u    is None: faltantes.append("UNIDAD")
        if col_lst  is None: faltantes.append("LISTA")
        if col_cont is None: faltantes.append("CONTADO")

        if faltantes:
            alertas.append(f"  [M2] Hoja '{sh}': no se encontraron columnas {faltantes}. "
                           f"Encabezados disponibles: {[str(h or '') for h in headers]}")
            # Intentar continuar con las que sí encontramos
            if col_u is None:
                continue

        for row in rows[header_idx + 1:]:
            if not row or len(row) <= col_u:
                continue
            unidad = str(row[col_u] or "").strip().upper()
            if not unidad or unidad in ("UNIDAD", "TIPO", "NONE", ""):
                continue
            lista = parse_price(row[col_lst])  if col_lst  is not None and col_lst  < len(row) else None
            cont  = parse_price(row[col_cont]) if col_cont is not None and col_cont < len(row) else None
            if lista or cont:
                resultado[unidad] = {"lista": lista, "cont": cont, "hoja": sh}

    return resultado, alertas


# ─── LECTURA OBRING ───────────────────────────────────────────────────────────
def read_obring(filepath):
    """
    Lee la hoja VALORES de la lista OBRING.
    Devuelve dict: {UNIDAD_UPPER: {"lista": float, "cont": float, "sup_cub": float, "sup_total": float}}
    Detecta columnas dinámicamente.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    alertas = []

    # Buscar hoja VALORES (o similar)
    hoja_vals = None
    for sh in wb.sheetnames:
        if "VALOR" in sh.upper():
            hoja_vals = sh
            break
    if hoja_vals is None:
        hoja_vals = wb.sheetnames[0]
        alertas.append(f"  [OBRING] No se encontró hoja 'VALORES'. Usando primera hoja: '{hoja_vals}'")

    ws = wb[hoja_vals]
    rows = list(ws.iter_rows(values_only=True))
    resultado = {}

    # Puede haber múltiples bloques de tabla en la misma hoja
    i = 0
    while i < len(rows):
        row = rows[i]
        cells_u = [str(c or "").strip().upper() for c in row]

        # Detectar fila de encabezado
        if ("PISO" in cells_u or "UNIDAD" in cells_u) and ("VALOR" in cells_u or "PRECIO" in cells_u):
            col_u     = find_col(row, "UNIDAD", "PISO")
            col_v     = find_col(row, "VALOR", "PRECIO")
            col_cub   = find_col(row, "M2 CUB", "SUP CUB", "CUB")
            col_total = find_col(row, "M2 TOTAL", "SUP TOTAL", "TOTAL")

            faltantes = []
            if col_u is None: faltantes.append("UNIDAD/PISO")
            if col_v is None: faltantes.append("VALOR/PRECIO")
            if faltantes:
                alertas.append(f"  [OBRING] Bloque en fila {i+1}: faltan columnas {faltantes}. "
                               f"Encabezados: {[str(c or '') for c in row]}")
                i += 1
                continue

            # Leer filas del bloque
            j = i + 1
            while j < len(rows) and j < i + 60:
                r = rows[j]
                if not r or len(r) <= col_u:
                    j += 1
                    continue
                unidad = str(r[col_u] or "").strip().upper()
                if not unidad:
                    j += 1
                    continue
                # Si encontramos otro encabezado, salir del bloque
                cells2 = [str(c or "").strip().upper() for c in r]
                if "PISO" in cells2 or "UNIDAD" in cells2:
                    break
                valor     = parse_price(r[col_v])     if col_v     < len(r) else None
                sup_cub   = parse_price(r[col_cub])   if col_cub   is not None and col_cub   < len(r) else None
                sup_total = parse_price(r[col_total])  if col_total is not None and col_total < len(r) else None
                if valor:
                    resultado[unidad] = {
                        "lista": valor, "cont": valor,
                        "sup_cub": sup_cub, "sup_total": sup_total
                    }
                j += 1
            i = j
        else:
            i += 1

    return resultado, alertas


# ─── LECTURA DISPONIBLES ──────────────────────────────────────────────────────
def read_disponibles():
    """
    Lee DISPONIBLES.xlsx. Devuelve lista de dicts con datos + numero de fila (1-based).
    """
    wb = openpyxl.load_workbook(DISP_FILE)
    ws = wb.active
    props = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 9:
            continue
        empresa = str(row[C_EMPRESA] or "").strip().upper()
        unidad  = str(row[C_UNIDAD]  or "").strip().upper()
        if not unidad:
            continue
        props.append({
            "fila":         i,
            "empresa":      empresa,
            "zona":         str(row[C_ZONA]       or "").strip(),
            "ubicacion":    str(row[C_UBICACION]  or "").strip(),
            "unidad":       unidad,
            "tipo":         str(row[C_TIPO]       or "").strip(),
            "precio_lista": parse_price(row[C_PRECIO_LISTA]) if len(row) > C_PRECIO_LISTA else None,
            "precio_cont":  parse_price(row[C_PRECIO_CONT])  if len(row) > C_PRECIO_CONT  else None,
            "usd_m2":       parse_price(row[C_USD_M2])        if len(row) > C_USD_M2       else None,
            "sup_cub":      parse_price(row[C_SUP_CUB])       if len(row) > C_SUP_CUB      else None,
            "sup_total":    parse_price(row[C_SUP_TOTAL])     if len(row) > C_SUP_TOTAL    else None,
        })
    return props


# ─── EXCLUSIONES ─────────────────────────────────────────────────────────────
def cargar_exclusiones():
    """
    Lee excluir.json. Estructura: {"M2": ["UNIDAD1", ...], "OBRING": ["UNIDAD2", ...]}
    Devuelve dict vacío si no existe.
    """
    if not os.path.exists(EXCLUIR_FILE):
        return {"M2": [], "OBRING": []}
    try:
        with open(EXCLUIR_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return {"M2": data.get("M2", []), "OBRING": data.get("OBRING", [])}
    except Exception as e:
        print(f"  AVISO: No se pudo leer excluir.json: {e}")
        return {"M2": [], "OBRING": []}


def guardar_exclusiones(nuevas_excluidas, excluidos_actuales):
    """Agrega las unidades marcadas con X al archivo excluir.json."""
    for n in nuevas_excluidas:
        emp = n["empresa"]
        u   = n["unidad"]
        if emp not in excluidos_actuales:
            excluidos_actuales[emp] = []
        if u not in excluidos_actuales[emp]:
            excluidos_actuales[emp].append(u)
    with open(EXCLUIR_FILE, "w", encoding="utf-8") as f:
        json.dump(excluidos_actuales, f, ensure_ascii=False, indent=2)


# ─── SINCRONIZACIÓN ───────────────────────────────────────────────────────────
def fmt_precio(v):
    if v is None:
        return "—"
    return f"U$S {int(v):,}".replace(",", ".")


def sincronizar():
    sep = "=" * 60

    print(f"\n{sep}")
    print("  SINCRONIZAR DISPONIBLES CON LISTAS NUEVAS")
    print(sep)

    if not os.path.exists(DISP_FILE):
        print(f"\n  ERROR: No se encontro {DISP_FILE}")
        return

    # ── Encontrar archivos de listas ──────────────────────────────────────────
    m2_file     = find_latest_file(M2_LISTAS_DIR)
    obring_file = find_latest_file(OBRING_DIR)

    print(f"\n  Lista M2:     {os.path.basename(m2_file)     if m2_file     else 'NO ENCONTRADA'}")
    print(f"  Lista OBRING: {os.path.basename(obring_file) if obring_file else 'NO ENCONTRADA'}")

    if not m2_file and not obring_file:
        print("\n  No se encontraron listas. Verifica las carpetas:")
        print(f"    {M2_LISTAS_DIR}")
        print(f"    {OBRING_DIR}")
        return

    # ── Leer listas nuevas ────────────────────────────────────────────────────
    m2_data     = {}
    obring_data = {}
    todas_alertas = []

    if m2_file:
        print(f"\n  Leyendo M2...")
        m2_data, alertas = read_m2(m2_file)
        todas_alertas += alertas
        print(f"  {len(m2_data)} unidades en lista M2")

    if obring_file:
        print(f"\n  Leyendo OBRING...")
        obring_data, alertas = read_obring(obring_file)
        todas_alertas += alertas
        print(f"  {len(obring_data)} unidades en lista OBRING")

    if todas_alertas:
        print(f"\n  ALERTAS DE FORMATO:")
        for a in todas_alertas:
            print(a)

    # ── Leer disponibles actuales ─────────────────────────────────────────────
    print(f"\n  Leyendo DISPONIBLES.xlsx...")
    props = read_disponibles()
    print(f"  {len(props)} propiedades actuales")

    # ── Comparar ──────────────────────────────────────────────────────────────
    cambios_precio = []   # (prop, precio_viejo, precio_nuevo)
    desaparecidos  = []   # props que ya no están en la lista correspondiente
    nuevos_m2      = []   # unidades en M2 que no están en DISPONIBLES
    nuevos_obring  = []   # unidades en OBRING que no están en DISPONIBLES

    # Keys de disponibles por empresa
    m2_en_disp     = {p["unidad"]: p for p in props if p["empresa"] == "M2"}
    obring_en_disp = {p["unidad"]: p for p in props if p["empresa"] == "OBRING"}

    # Cambios de precio y desaparecidos M2
    for p in props:
        if p["empresa"] != "M2":
            continue
        u = p["unidad"]
        if u in m2_data:
            nuevo_cont = m2_data[u]["cont"]
            nuevo_list = m2_data[u]["lista"]
            if nuevo_cont and nuevo_cont != p["precio_cont"]:
                cambios_precio.append({
                    "prop": p,
                    "lista_vieja": p["precio_lista"],
                    "cont_vieja":  p["precio_cont"],
                    "lista_nueva": nuevo_list,
                    "cont_nueva":  nuevo_cont,
                })
        else:
            desaparecidos.append(p)

    # Cambios de precio y desaparecidos OBRING
    for p in props:
        if p["empresa"] != "OBRING":
            continue
        u = p["unidad"]
        if u in obring_data:
            nuevo_cont = obring_data[u]["cont"]
            nuevo_list = obring_data[u]["lista"]
            if nuevo_cont and nuevo_cont != p["precio_cont"]:
                cambios_precio.append({
                    "prop": p,
                    "lista_vieja": p["precio_lista"],
                    "cont_vieja":  p["precio_cont"],
                    "lista_nueva": nuevo_list,
                    "cont_nueva":  nuevo_cont,
                })
        else:
            desaparecidos.append(p)

    # ── Cargar exclusiones ────────────────────────────────────────────────────
    excluidos = cargar_exclusiones()
    excluidos_m2     = excluidos.get("M2", [])
    excluidos_obring = excluidos.get("OBRING", [])

    # Unidades nuevas (en lista pero no en DISPONIBLES, filtrando excluidas)
    for u, d in m2_data.items():
        if u not in m2_en_disp and u not in excluidos_m2:
            nuevos_m2.append({"unidad": u, "hoja": d.get("hoja",""), "cont": d["cont"], "lista": d["lista"]})

    for u, d in obring_data.items():
        if u not in obring_en_disp and u not in excluidos_obring:
            nuevos_obring.append({"unidad": u, "cont": d["cont"], "lista": d["lista"]})

    # ── Mostrar resumen ───────────────────────────────────────────────────────
    print(f"\n{sep}")
    print("  RESUMEN DE CAMBIOS")
    print(sep)
    print(f"\n  Cambios de precio:        {len(cambios_precio)}")
    print(f"  No estan en lista nueva:  {len(desaparecidos)}")
    print(f"  Unidades nuevas en M2:    {len(nuevos_m2)}")
    print(f"  Unidades nuevas en OBRING:{len(nuevos_obring)}")
    if excluidos_m2 or excluidos_obring:
        tot = len(excluidos_m2) + len(excluidos_obring)
        print(f"  (Ignoradas por exclusion: {tot} unidades)")

    if not cambios_precio and not desaparecidos and not nuevos_m2 and not nuevos_obring:
        print("\n  Todo actualizado. No hay cambios.")
        return

    # ── Detalle de cambios de precio ──────────────────────────────────────────
    if cambios_precio:
        print(f"\n  CAMBIOS DE PRECIO ({len(cambios_precio)} unidades):")
        print(f"  {'Empresa':<8} {'Unidad':<12} {'Ub./Proyecto':<20} {'Precio ant.':<14} {'Precio nuevo':<14}")
        print(f"  {'-'*8} {'-'*12} {'-'*20} {'-'*14} {'-'*14}")
        for c in cambios_precio:
            p = c["prop"]
            ref = p["ubicacion"][:20]
            print(f"  {p['empresa']:<8} {p['unidad']:<12} {ref:<20} "
                  f"{fmt_precio(c['cont_vieja']):<14} {fmt_precio(c['cont_nueva']):<14}")

    # ── Detalle de desaparecidos ──────────────────────────────────────────────
    if desaparecidos:
        print(f"\n  DESAPARECIDOS DE LA LISTA ({len(desaparecidos)} unidades):")
        print(f"  {'#':<4} {'Empresa':<8} {'Unidad':<12} {'Ub./Proyecto':<25} {'Precio cont.'}")
        print(f"  {'-'*4} {'-'*8} {'-'*12} {'-'*25} {'-'*14}")
        for i, p in enumerate(desaparecidos, 1):
            ref = p["ubicacion"][:25]
            print(f"  {i:<4} {p['empresa']:<8} {p['unidad']:<12} {ref:<25} {fmt_precio(p['precio_cont'])}")

    # ── Nuevas unidades: decidir ahora mismo ─────────────────────────────────
    agregar_nuevos = []
    excluir_nuevos = []
    todos_nuevos = (
        [dict(n, empresa="M2")     for n in nuevos_m2] +
        [dict(n, empresa="OBRING") for n in nuevos_obring]
    )
    if todos_nuevos:
        print(f"\n  NUEVAS UNIDADES DETECTADAS ({len(todos_nuevos)}):")
        print(f"  A = Agregar  |  X = Excluir siempre  |  Enter = Ignorar por ahora")
        print()
        for n in todos_nuevos:
            emp  = n["empresa"]
            hoja = f" [{n['hoja']}]" if n.get("hoja") else ""
            print(f"  {emp}{hoja} | {n['unidad']:<20} | {fmt_precio(n['cont'])}")
            dec = input("  [A/X/Enter]: ").strip().upper()
            if dec == "A":
                agregar_nuevos.append(n)
            elif dec == "X":
                excluir_nuevos.append(n)

    # ── Confirmación ─────────────────────────────────────────────────────────
    print(f"\n{sep}")

    # Aplicar cambios de precio
    aplicar_precios = False
    if cambios_precio:
        r = input(f"\n  Actualizar {len(cambios_precio)} precios en DISPONIBLES.xlsx? [s/N]: ").strip().lower()
        aplicar_precios = (r == "s")

    # Desaparecidos: mostrar y preguntar cuáles eliminar
    eliminar_filas = []
    if desaparecidos:
        print(f"\n  NO ESTAN EN LA LISTA NUEVA ({len(desaparecidos)} unidades):")
        print(f"  {'#':<4} {'Empresa':<8} {'Unidad':<14} {'Ub./Proyecto':<25} {'Precio cont.'}")
        print(f"  {'-'*4} {'-'*8} {'-'*14} {'-'*25} {'-'*14}")
        for i, p in enumerate(desaparecidos, 1):
            ref = p["ubicacion"][:25]
            print(f"  {i:<4} {p['empresa']:<8} {p['unidad']:<14} {ref:<25} {fmt_precio(p['precio_cont'])}")
        print(f"\n  Numeros a ELIMINAR separados por coma (ej: 1,3,5),")
        print(f"  'todos' para eliminar todas, Enter para no eliminar ninguna.")
        r = input("  > ").strip().lower()
        if r == "todos":
            eliminar_filas = [p["fila"] for p in desaparecidos]
        elif r:
            try:
                nums = [int(x.strip()) for x in r.split(",")]
                eliminar_filas = [desaparecidos[n-1]["fila"] for n in nums if 1 <= n <= len(desaparecidos)]
            except Exception:
                print("  Entrada invalida, no se elimina nada.")

    hay_cambios = aplicar_precios or eliminar_filas or agregar_nuevos or excluir_nuevos
    if not hay_cambios:
        print("\n  Sin cambios aplicados.")
        return

    # ── Guardar exclusiones permanentes ──────────────────────────────────────
    if excluir_nuevos:
        guardar_exclusiones(excluir_nuevos, excluidos)
        print(f"  ✓ {len(excluir_nuevos)} unidad(es) agregada(s) a la lista de exclusion")

    if not aplicar_precios and not eliminar_filas and not agregar_nuevos:
        return

    # ── Aplicar al xlsx ───────────────────────────────────────────────────────
    print(f"\n  Aplicando cambios...")
    wb = openpyxl.load_workbook(DISP_FILE)
    ws = wb.active

    # Actualizar precios
    if aplicar_precios:
        cambio_map = {c["prop"]["fila"]: c for c in cambios_precio}
        for row in ws.iter_rows(min_row=2):
            fila = row[0].row
            if fila not in cambio_map:
                continue
            c = cambio_map[fila]
            row[C_PRECIO_LISTA].value = c["lista_nueva"]
            row[C_PRECIO_CONT].value  = c["cont_nueva"]
            sup = parse_price(row[C_SUP_TOTAL].value) or parse_price(row[C_SUP_CUB].value)
            if sup and c["cont_nueva"]:
                row[C_USD_M2].value = round(c["cont_nueva"] / sup, 2)
        print(f"  ✓ {len(cambios_precio)} precios actualizados")

    # Agregar nuevas unidades (fila vacía con datos básicos al final)
    if agregar_nuevos:
        for n in agregar_nuevos:
            # Fila mínima: empresa en col 1, unidad en col 8, precios en 16/17
            nueva_fila = [""] * 22
            nueva_fila[C_EMPRESA]      = n["empresa"]
            nueva_fila[C_UNIDAD]       = n["unidad"]
            nueva_fila[C_PRECIO_LISTA] = n.get("lista")
            nueva_fila[C_PRECIO_CONT]  = n.get("cont")
            ws.append(nueva_fila)
        print(f"  ✓ {len(agregar_nuevos)} unidad(es) agregada(s) al xlsx")
        print(f"    IMPORTANTE: completar zona, direccion, tipo, etc. en DISPONIBLES.xlsx")

    # Eliminar filas (de atrás para adelante)
    if eliminar_filas:
        _guardar_historial(props, eliminar_filas, "sincronizacion")
        for fila in sorted(eliminar_filas, reverse=True):
            ws.delete_rows(fila)
        print(f"  ✓ {len(eliminar_filas)} unidades eliminadas")

    # Guardar
    ts  = datetime.now().strftime("%Y%m%d_%H%M")
    bak = DISP_FILE.replace(".xlsx", f"_bak_{ts}.xlsx")
    import shutil
    shutil.copy2(DISP_FILE, bak)
    wb.save(DISP_FILE)
    print(f"  ✓ Guardado. Backup: {os.path.basename(bak)}")

    print(f"\n  Ahora ejecutá opción 2 para regenerar el HTML.")
    print(sep)


# ─── HISTORIAL ────────────────────────────────────────────────────────────────
def _guardar_historial(props, filas_eliminadas, motivo):
    import json
    historial_file = os.path.join(DISPONIBLES_DIR, "vendidas.json")
    historial = []
    if os.path.exists(historial_file):
        try:
            with open(historial_file, encoding="utf-8") as f:
                historial = json.load(f)
        except Exception:
            pass

    fila_set = set(filas_eliminadas)
    ts = datetime.now().isoformat(timespec="seconds")
    for p in props:
        if p["fila"] in fila_set:
            historial.append({
                "fecha_baja": ts,
                "motivo":     motivo,
                "empresa":    p["empresa"],
                "unidad":     p["unidad"],
                "ubicacion":  p.get("ubicacion",""),
                "zona":       p.get("zona",""),
                "tipo":       p.get("tipo",""),
                "precio_cont": p.get("precio_cont"),
            })

    with open(historial_file, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)
    print(f"  ✓ Historial guardado: {len(filas_eliminadas)} entrada(s) en vendidas.json")


if __name__ == "__main__":
    try:
        sincronizar()
    except Exception as e:
        import traceback
        print(f"\n  ERROR INESPERADO: {e}")
        traceback.print_exc()
