"""
Actualizar DISPONIBLES.xlsx con precios de listas M2 y Obring
Uso: python actualizar_disponibles.py
Los archivos Excel se buscan en sus carpetas originales dentro de DESARROLLOS.
"""

import openpyxl
from datetime import datetime
import os

# Este script vive en: DESARROLLOS/DISPONIBLES/
DISPONIBLES_DIR = os.path.dirname(os.path.abspath(__file__))
DESARROLLOS_DIR = os.path.dirname(DISPONIBLES_DIR)

DISP_FILE   = os.path.join(DISPONIBLES_DIR, "DISPONIBLES.xlsx")
M2_FILE     = os.path.join(DESARROLLOS_DIR, "M2 DESARROLLOS", "LISTAS DE PRECIO", "Lista de precios Mayo 2026.xlsx")
OBRING_FILE = os.path.join(DESARROLLOS_DIR, "OBRING", "OBRING MAYO 2026.xlsx")

# Columnas de DISPONIBLES (0-based, sin columna PROYECTO)
COL_EMPRESA      = 1
COL_UNIDAD       = 7
COL_PRECIO_LISTA = 15
COL_PRECIO_CONT  = 16
COL_USD_M2       = 17
COL_SUP_CUB      = 10
COL_SUP_TOTAL    = 14


def parse_price(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("USD", "").replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except Exception:
        return None


def extract_m2_prices(filepath):
    """Lee todas las hojas de M2 y devuelve dict {unidad_upper: (lista, contado)}."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    prices = {}
    for shname in wb.sheetnames:
        ws = wb[shname]
        header_row = None
        col_unidad = col_lista = col_cont = None
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            cells = [str(c or "").strip().upper() for c in row]
            if "UNIDAD" in cells:
                header_row = i
                col_unidad = cells.index("UNIDAD")
                for j, c in enumerate(cells):
                    if "LISTA" in c:
                        col_lista = j
                    if "CONTADO" in c:
                        col_cont = j
                break
        if header_row is None or col_unidad is None:
            continue
        for row in rows[header_row + 1:]:
            if len(row) <= col_unidad:
                continue
            unidad = str(row[col_unidad] or "").strip()
            if not unidad or unidad.lower() in ("unidad", "none", ""):
                continue
            lista = parse_price(row[col_lista]) if col_lista is not None and col_lista < len(row) else None
            cont  = parse_price(row[col_cont])  if col_cont  is not None and col_cont  < len(row) else None
            if lista or cont:
                prices[unidad.upper()] = (lista, cont)
    return prices


def extract_obring_prices(filepath):
    """Lee OBRING VALORES y devuelve dict {unidad_upper: (valor, valor, sup_cub, sup_total)}."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["VALORES"]
    prices = {}
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        cells = [str(c or "").strip().upper() for c in row]
        if ("PISO" in cells or "UNIDAD" in cells) and "VALOR" in cells:
            try:
                col_u = cells.index("UNIDAD") if "UNIDAD" in cells else cells.index("PISO")
            except Exception:
                continue
            col_v     = cells.index("VALOR")
            col_cub   = cells.index("M2 CUB")   if "M2 CUB"   in cells else None
            col_total = cells.index("M2 TOTAL")  if "M2 TOTAL" in cells else None
            for j in range(i + 1, min(i + 30, len(rows))):
                r      = rows[j]
                cells2 = [str(c or "").strip() for c in r]
                unidad = cells2[col_u] if col_u < len(cells2) else ""
                if not unidad:
                    continue
                valor     = parse_price(r[col_v])     if col_v     < len(r) else None
                sup_cub   = parse_price(r[col_cub])   if col_cub   is not None and col_cub   < len(r) else None
                sup_total = parse_price(r[col_total])  if col_total is not None and col_total < len(r) else None
                if valor:
                    prices[unidad.upper()] = (valor, valor, sup_cub, sup_total)
    return prices


def actualizar():
    # Buscar la lista de precios M2 mas reciente si la de Mayo no existe
    m2_path = M2_FILE
    if not os.path.exists(m2_path):
        listas_dir = os.path.join(DESARROLLOS_DIR, "M2 DESARROLLOS", "LISTAS DE PRECIO")
        if os.path.exists(listas_dir):
            xlsx_files = sorted([f for f in os.listdir(listas_dir) if f.endswith('.xlsx')], reverse=True)
            if xlsx_files:
                m2_path = os.path.join(listas_dir, xlsx_files[0])
                print(f"Lista M2 no encontrada, usando: {xlsx_files[0]}")

    print(f"Leyendo {DISP_FILE} ...")
    if not os.path.exists(DISP_FILE):
        print(f"ERROR: No se encontro {DISP_FILE}")
        return

    wb_disp = openpyxl.load_workbook(DISP_FILE)
    ws = wb_disp.active

    print("Leyendo precios M2 ...")
    if os.path.exists(m2_path):
        m2_prices = extract_m2_prices(m2_path)
        print(f"  {len(m2_prices)} unidades")
    else:
        print(f"  AVISO: No se encontro {m2_path}")
        m2_prices = {}

    print("Leyendo precios OBRING ...")
    if os.path.exists(OBRING_FILE):
        obring_prices = extract_obring_prices(OBRING_FILE)
        print(f"  {len(obring_prices)} unidades")
    else:
        print(f"  AVISO: No se encontro {OBRING_FILE}")
        obring_prices = {}

    actualizados = 0
    for row in ws.iter_rows(min_row=2):
        empresa = str(row[COL_EMPRESA].value or "").strip().upper()
        unidad  = str(row[COL_UNIDAD].value  or "").strip().upper()
        if not unidad:
            continue

        nueva_lista = nueva_cont = None

        if empresa == "M2" and unidad in m2_prices:
            nueva_lista, nueva_cont = m2_prices[unidad]
        elif empresa == "OBRING" and unidad in obring_prices:
            nueva_lista, nueva_cont, sup_cub, sup_total = obring_prices[unidad]
            if sup_cub and not row[COL_SUP_CUB].value:
                row[COL_SUP_CUB].value = sup_cub
            if sup_total and not row[COL_SUP_TOTAL].value:
                row[COL_SUP_TOTAL].value = sup_total

        if nueva_lista or nueva_cont:
            viejo_cont = parse_price(row[COL_PRECIO_CONT].value)
            if viejo_cont != nueva_cont:
                row[COL_PRECIO_LISTA].value = nueva_lista
                row[COL_PRECIO_CONT].value  = nueva_cont
                sup = parse_price(row[COL_SUP_TOTAL].value) or parse_price(row[COL_SUP_CUB].value)
                if sup and nueva_cont:
                    row[COL_USD_M2].value = round(nueva_cont / sup, 2)
                actualizados += 1
                print(f"  OK  {empresa} {unidad}: {viejo_cont} -> {nueva_cont}")

    import shutil
    ts  = datetime.now().strftime("%Y%m%d_%H%M")
    bak = os.path.join(DISPONIBLES_DIR, f"DISPONIBLES_bak_{ts}.xlsx")
    shutil.copy2(DISP_FILE, bak)
    wb_disp.save(DISP_FILE)
    print(f"\nActualizados: {actualizados} registros")
    print(f"Guardado en:  DISPONIBLES.xlsx (backup: {os.path.basename(bak)})")
    print("\nAhora ejecuta opcion 8 para regenerar y publicar.")


if __name__ == "__main__":
    actualizar()
    input("\nPresiona Enter para cerrar...")
